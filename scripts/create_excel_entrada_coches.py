from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


OUTPUT_DIR = Path(__file__).resolve().parent / "output"
OUTPUT_FILE = OUTPUT_DIR / "entrada_coches_nuevo_local.xlsx"


def build_template(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Entradas"

    # Identidad visual SW Studio (sin parecerse al entorno anterior)
    c_bg_header = "0F2A36"
    c_bg_title = "134455"
    c_bg_note = "EAF7FB"
    c_text_white = "FFFFFF"
    c_text_dark = "10212B"
    c_border = "BFD6DF"

    ws.merge_cells("A1:N1")
    ws["A1"] = "SW Studio · Registro de Entrada de Coches (Nuevo Local)"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=c_text_white)
    ws["A1"].fill = PatternFill("solid", fgColor=c_bg_title)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:N2")
    ws["A2"] = (
        "Uso diario: una fila por coche. Completa recepción, estado y entrega. "
        f"Plantilla generada: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )
    ws["A2"].font = Font(name="Calibri", size=10, color=c_text_dark)
    ws["A2"].fill = PatternFill("solid", fgColor=c_bg_note)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    headers = [
        "Fecha entrada",
        "Hora",
        "Matrícula",
        "Cliente",
        "Teléfono",
        "Vehículo",
        "KM entrada",
        "Servicio principal",
        "Estado",
        "Asesor",
        "Fecha entrega prevista",
        "Importe estimado (€)",
        "Método cobro",
        "Observaciones",
    ]

    ws.append([])
    ws.append(headers)

    header_row = 4
    thin = Side(style="thin", color=c_border)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = Font(name="Calibri", size=10, bold=True, color=c_text_white)
        cell.fill = PatternFill("solid", fgColor=c_bg_header)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # Filas iniciales vacias para trabajar
    start_data = 5
    end_data = 204
    for r in range(start_data, end_data + 1):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 14))

    # Validaciones visuales basicas
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A{header_row}:N{end_data}"

    # Tabla con estilo para filtros y lectura
    table = Table(displayName="EntradasCoches", ref=f"A{header_row}:N{end_data}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Anchos de columna pensados para operativa diaria
    widths = {
        "A": 14,
        "B": 9,
        "C": 13,
        "D": 24,
        "E": 14,
        "F": 22,
        "G": 11,
        "H": 22,
        "I": 13,
        "J": 14,
        "K": 18,
        "L": 16,
        "M": 14,
        "N": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Formatos de numero/fecha
    for r in range(start_data, end_data + 1):
        ws[f"A{r}"].number_format = "DD/MM/YYYY"
        ws[f"K{r}"].number_format = "DD/MM/YYYY"
        ws[f"L{r}"].number_format = "#,##0.00"

    # Hoja de resumen rapido
    rs = wb.create_sheet("Resumen")
    rs["A1"] = "Resumen diario"
    rs["A1"].font = Font(size=13, bold=True, color=c_text_white)
    rs["A1"].fill = PatternFill("solid", fgColor=c_bg_title)
    rs.merge_cells("A1:D1")

    rs["A3"] = "Total coches registrados"
    rs["B3"] = "=COUNTA(Entradas!C5:C204)"

    rs["A4"] = "Pendientes entrega"
    rs["B4"] = '=COUNTIFS(Entradas!I5:I204,"<>Entregado",Entradas!C5:C204,"<>")'

    rs["A5"] = "Entregados"
    rs["B5"] = '=COUNTIFS(Entradas!I5:I204,"Entregado")'

    rs["A6"] = "Importe estimado total (€)"
    rs["B6"] = "=SUM(Entradas!L5:L204)"

    rs["A8"] = "Estados sugeridos"
    rs["A9"] = "Recepcionado"
    rs["A10"] = "En proceso"
    rs["A11"] = "Pendiente pieza"
    rs["A12"] = "Listo"
    rs["A13"] = "Entregado"

    for row in range(3, 14):
        rs[f"A{row}"].font = Font(bold=(row in {3, 4, 5, 6, 8}))
    rs.column_dimensions["A"].width = 32
    rs.column_dimensions["B"].width = 18

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


if __name__ == "__main__":
    build_template(OUTPUT_FILE)
    print(f"Plantilla creada: {OUTPUT_FILE}")
